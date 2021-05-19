# chapter 42 and 43 are important.

# Question: Given one excel sheet, you need to print out all the values from it.

import pandas
import openpyxl

# for i in dir(openpyxl):
#     if not i.startswith('_'):
#         print(i)

book = openpyxl.load_workbook(
    r"C:\Users\Chandan Mukherjee\PycharmProjects\ProperProjectChandan\Automate_the_boring_stuff\Test_Excel_for_Automation.xlsx")
sheet_names = book.sheetnames
# print(sheet_names)
sheet_used = book['Sheet1']

# print(sheet_used['A1'].value)
# print(sheet_used.cell(row=1, column=1).value)

# alphs = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
# nums = list('123456789')
# coordinates=[]
# for i in alphs:
#     for j in nums:
#         coordinates.append(str(i+j))
# # print(coordinates)
#




# Writting excel sheets

# if you want to create new workbook, do not use load_workbook method, instead use workbook() method.


sheet_used['A1'].value= 500
sheet_used['B1'].value= 2000

book.save(r"C:\Users\Chandan Mukherjee\PycharmProjects\ProperProjectChandan\Automate_the_boring_stuff\Test_Excel_for_Automation.xlsx")
# you must use save() method to save the changes made in your file.

# create_sheet is used to create sheet

# sheet.title is used to change the title of the sheet.

